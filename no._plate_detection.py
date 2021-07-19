import cv2    # for image-processing
import datetime
import openpyxl    # for operation on excel
from openpyxl import load_workbook
import os
import pytesseract    # for image to text

set = []
count = 0
X = []
Y = []
cc = 1
state = None

cap = cv2.VideoCapture('q2.mp4')    # capture the video

while True:
    ret, img = cap.read()    # reading the video frame by frame
    lx = img.shape[1]
    ly = img.shape[0]
    cv2.line(img , (0,int(ly/1.5)) , (int(lx),int(ly/1.5)) , (0,0,255), 2)    # drawing the line
    plate=cv2.CascadeClassifier('indian_license_plate.xml')    #identifing the number plate
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    pt = plate.detectMultiScale(gray, 1.3, 5)
    
    
    now = datetime.datetime.now()
    day = now.day
    b = str(day) + '.' + str(now.month) + '.' + str(now.year)    # date .
    mainfile = b+'.xlsx'    # excel file name
    if not(os.path.isfile(mainfile)):
        mwb = openpyxl.Workbook()
        sheet = mwb.active
        a = 'Daily Vehicles'
        day = now.day
        b = str(str(a) + '-' + str(day) + '.' + str(now.month) + '.' + str(now.year))
        sheet.title = b
        sheet.cell(row=1, column=1).value = 'Time'
        sheet.cell(row=1, column=2).value = 'Detected no. (by pytesseract)'
        sheet.cell(row=1, column=3).value = 'Number Plate'
        sheet.cell(row=1, column=4).value = 'IN/OUT'
        mwb.save(mainfile)
    if len(pt) >= 2:
        pt = [pt[len(pt)-1]]
    else:
        pt = pt
    for (x,y,w,h) in pt:
        cv2.rectangle(img, (x, y), (x + w, y + h), (255,0,0),2)    # puting the rectangle around the number plate
        cv2.putText(img, "Status: {}".format('plate detected'), (30, 60),
                    cv2.FONT_HERSHEY_SIMPLEX, 2, (0, 0, 255), 3)
        plate = img[y:y+h, x:x+w]    # Number plate image
        X.append(x)
        Y.append(y)
        count = count + 1
        if count == 5:
            if Y[0]<int(ly/1.5)<Y[4] and abs(Y[0] - Y[4]) <= 100:
                cc = cc + 1
                cv2.imwrite("output_plate_image/plate"+str(cc-1)+".jpg", img[y:y+h, x:x+w])    # Number plate image saved
                state = "IN"
                
                c = datetime.datetime.now()
                e = c.strftime('%H') + ':' + c.strftime('%M') + ':' + c.strftime('%S')    # time
                
                plateimg = "output_plate_image/plate"+str(cc-1)+".jpg"
                pic = openpyxl.drawing.image.Image(plateimg)
                
                img1 = cv2.imread("output_plate_image/plate"+str(cc-1)+".jpg")
                s = ''
                gray = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY)
                blur = cv2.bilateralFilter(gray, 11,90, 90)
                edges = cv2.Canny(blur, 30, 200)
                cnts, new = cv2.findContours(edges.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
                for c in cnts:
                    x,y,w,h = cv2.boundingRect(c)
                    no = img1[y:y+h, x:x+w]    # Image of the letters or numbers in the number plate
                    text = pytesseract.image_to_string(no, lang="eng")    # Image to text (string)
                    s = s + text    # Number on the number plate in the string formate
                
                mwb = load_workbook(mainfile)
                sheet = mwb.active
                sheet.cell(row=cc, column=1).value = e
                sheet.cell(row=cc, column=2).value = s
                sheet.cell(row=cc, column=4).value = state
                sheet.add_image(pic, 'C%d' % cc)
                sheet.row_dimensions[cc].height = 50
                sheet.column_dimensions['A'].width = 20
                sheet.column_dimensions['B'].width = 30
                sheet.column_dimensions['C'].width = 40
                sheet.column_dimensions['D'].width = 20
                mwb.save(mainfile)
                
            elif Y[0]>int(ly/3)>Y[4] and abs(Y[0] - Y[4]) <= 100:               
                cc = cc + 1
                cv2.imwrite("output_plate_image/plate"+str(cc-1)+".jpg", img[y:y+h, x:x+w])
                state = "OUT"
                
                c = datetime.datetime.now()
                e = c.strftime('%H') + ':' + c.strftime('%M') + ':' + c.strftime('%S')    # time
                
                plateimg = "output_plate_image/plate"+str(cc-1)+".jpg"
                pic = openpyxl.drawing.image.Image(plateimg)
                
                img1 = cv2.imread("output_plate_image/plate"+str(cc-1)+".jpg")
                s = ''
                gray = cv2.cvtColor(img1, cv2.COLOR_BGR2GRAY)
                blur = cv2.bilateralFilter(gray, 11,90, 90)
                edges = cv2.Canny(blur, 30, 200)
                cnts, new = cv2.findContours(edges.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
                for c in cnts:
                    x,y,w,h = cv2.boundingRect(c)
                    no = img1[y:y+h, x:x+w]    # Image of the letters or numbers in the number plate
                    text = pytesseract.image_to_string(no, lang="eng")    # Image to text (string)
                    s = s + text    # Number on the number plate in the string formate
                
                mwb = load_workbook(mainfile)
                sheet = mwb.active
                sheet.cell(row=cc, column=1).value = e
                sheet.cell(row=cc, column=2).value = s
                sheet.cell(row=cc, column=4).value = state
                sheet.add_image(pic, 'C%d' % cc)
                sheet.row_dimensions[cc].height = 50
                qsheet.column_dimensions['A'].width = 20    # Setting the size of the excel cell
                sheet.column_dimensions['B'].width = 30    # Setting the size of the excel cell
                sheet.column_dimensions['C'].width = 40    # Setting the size of the excel cell
                sheet.column_dimensions['D'].width = 20    # Setting the size of the excel cell
                mwb.save(mainfile)
                
            count = 0
            Y = []
            X = []
            
    cv2.imshow('Capturing Video',img)
    if(cv2.waitKey(1) & 0xFF == ord('q')):
        cap.release()
        cv2.destroyAllWindows()
